from flask import Flask, render_template, redirect, url_for, flash, request
import openpyxl
from openpyxl.styles import Font
import os

app = Flask(__name__)
app.secret_key = 'LandingPage' #necessary for using flash messages

@app.route('/', methods=['GET', 'POST'])
def home():
    
    # POST method to handle form data
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        email = request.form.get('email')

        errors = False
        # Validation of data entry
        if name == '' or not all(char.isalpha() or char.isspace() for char in name):
            flash('Please enter valid name (Only letters allowed)', 'error-name')
            errors = True
        
        if phone == '' or not phone.isdigit():
            flash('Please enter valid contact number (Only numbers allowed)', 'error-phone')
            errors = True

        if email == '':
            flash('Please enter valid email address', 'error-email')
            errors = True

        if errors:
            return render_template('index.html')

        # Excel file
        clients_list = 'clients.xlsx'

        # If Excel file doesn't exist, creates it and add headers
        if not os.path.exists(clients_list):
            wb = openpyxl.Workbook()
            sheet = wb.active

            headers = ['ID', 'Full Name', 'Contact Number', 'Email']
            bold_font = Font(bold=True)

            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = bold_font
        
        # If it exists, open it and add new data from the form
        else:
            wb = openpyxl.load_workbook(clients_list)
            sheet = wb.active

        next_row = sheet.max_row + 1

        last_id_cell = sheet.cell(row=next_row - 1, column=1).value

        if next_row == 2:
            new_id = 1

        else:
            new_id = last_id_cell + 1

        new_client = [new_id, name, phone, email]

        for col, val in enumerate(new_client, start=1):
            sheet.cell(row=next_row, column=col, value=val)

        wb.save(clients_list)

        flash('Form submitted successfully', 'form-success')

        return redirect(url_for('home'))

    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/treatment')
def treatment():
    return render_template('treatment.html')

@app.route('/contact')
def contact():
    return render_template('contact.html')

if __name__ == '__main__':
    app.run(debug=True)